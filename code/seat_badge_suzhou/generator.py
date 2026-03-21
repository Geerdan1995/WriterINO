# 导入操作系统相关功能模块（文件/目录操作等）
import os
# 导入系统相关功能模块（命令行参数、路径等）
import sys
# 导入与C语言兼容的类型模块（可能用于系统级操作）
import ctypes
# 导入tkinter库，用于创建GUI界面
import tkinter as tk
# 从tkinter导入文件对话框、消息框、带滚动条的文本框组件
from tkinter import filedialog, messagebox, scrolledtext
# 从tkinter导入ttk模块（提供更美观的界面组件）
from tkinter import ttk
# 导入pandas库，用于处理Excel表格数据
import pandas as pd
# 从openpyxl导入load_workbook，用于读取和写入Excel文件
from openpyxl import load_workbook
# 从pypinyin导入拼音转换相关功能
from pypinyin import pinyin, Style
# 从reportlab导入pdfgen模块，用于生成PDF文件
from reportlab.pdfgen import canvas
# 从reportlab导入页面大小相关常量（A4纸、纵向等）
from reportlab.lib.pagesizes import A4, portrait
# 从reportlab导入图片处理工具
from reportlab.lib.utils import ImageReader
# 从reportlab导入PDF字体相关模块
from reportlab.pdfbase import pdfmetrics
# 从reportlab导入TrueType字体处理类
from reportlab.pdfbase.ttfonts import TTFont
# 从datetime导入datetime类，用于处理时间
from datetime import datetime
# 导入idlelib（可能用于备用的界面支持，此处未直接使用）
import idlelib


# ========== 纯业务逻辑函数（可被 Web 后端调用） ==========

def setup_pdf_fonts(fonts_folder):
    """配置PDF中文字体"""
    try:
        font_file = "msyh.ttc"
        font_bold_file = "msyhbd.ttc"
        
        font_path = os.path.join(fonts_folder, font_file)
        font_bold_path = os.path.join(fonts_folder, font_bold_file)
        
        if not os.path.exists(font_path):
            sys_font_path = os.path.join("C:", "Windows", "Fonts")
            font_path = os.path.join(sys_font_path, font_file)
        if not os.path.exists(font_bold_path):
            sys_font_path = os.path.join("C:", "Windows", "Fonts")
            font_bold_path = os.path.join(sys_font_path, font_bold_file)
        
        pdfmetrics.registerFont(TTFont("Chinese", font_path))
        pdfmetrics.registerFont(TTFont("Chinese-Bold", font_bold_path))
        return True
    except Exception as e:
        print(f"字体注册警告: {str(e)}")
        return False


def find_excel_file(folder):
    """查找文件夹中的Excel文件（确保只有一个）"""
    excel_files = [os.path.join(folder, f) for f in os.listdir(folder)
                  if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    
    if len(excel_files) != 1:
        raise Exception(f"找到{len(excel_files)}个Excel文件，需且只能有1个")
    
    return excel_files[0]


def extract_department(org_path, department_list=None):
    """从组织全路径中提取部门信息"""
    if not isinstance(org_path, str):
        return ""
    
    if org_path.startswith("汇川集团"):
        parts = org_path.split("/")
        if "总部" in parts:
            总部索引 = parts.index("总部")
            if 总部索引 + 1 < len(parts):
                return parts[总部索引 + 1]
    else:
        if "/" in org_path:
            return org_path.split("/")[0]
    
    return org_path


def process_data_in_memory(excel_path, department_list=None):
    """
    在内存中处理Excel数据，生成员工信息列表（不生成中间文件）
    
    参数:
        excel_path: 原始Excel文件路径
        department_list: 预设部门列表（可选）
    
    返回:
        员工信息列表（dict格式）
    """
    source_df = pd.read_excel(excel_path)
    required_cols = ["姓名", "预入职工号", "岗位", "组织全路径"]
    for col in required_cols:
        if col not in source_df.columns:
            raise Exception(f"原始Excel缺少列: {col}")
    
    employees = []
    for _, row in source_df.iterrows():
        name = row["姓名"]
        emp_id = row["预入职工号"]
        position = row["岗位"]
        org_path = row["组织全路径"]
        
        department = extract_department(org_path, department_list)
        
        # 生成拼音
        pinyin_text = ""
        if name:
            surname = name[0]
            given_name = name[1:]
            surname_pinyin = pinyin(surname, style=Style.NORMAL)[0][0].capitalize()
            given_pinyin = ""
            if given_name:
                given_parts = pinyin(given_name, style=Style.NORMAL)
                given_pinyin = given_parts[0][0].capitalize() + ''.join([p[0] for p in given_parts[1:]])
            pinyin_text = f"{surname_pinyin} {given_pinyin}".strip()
        
        # 生成照片文件名
        photo_name = f"{emp_id}.jpg" if emp_id else ""
        
        employees.append({
            "姓名": name,
            "工号": emp_id,
            "岗位": position,
            "部门": department,
            "拼音": pinyin_text,
            "照片文件命名": photo_name
        })
    
    return employees


def generate_pdf_from_employees(employees, photo_folder, assets_folder, output_folder, fonts_folder):
    """
    从员工信息列表生成PDF（不依赖中间Excel文件）
    
    参数:
        employees: 员工信息列表
        photo_folder: 员工照片文件夹路径
        assets_folder: 资源文件夹路径
        output_folder: 输出文件夹路径
        fonts_folder: 字体文件夹路径
    
    返回:
        生成的PDF文件路径
    """
    setup_pdf_fonts(fonts_folder)
    
    top_img = os.path.join(assets_folder, "top.jpg")
    bot_img = os.path.join(assets_folder, "bot.jpg")
    default_photo = os.path.join(assets_folder, "bird.jpg")
    
    if not os.path.exists(top_img):
        raise Exception(f"未找到top.jpg")
    if not os.path.exists(bot_img):
        raise Exception(f"未找到bot.jpg")
    
    if not employees:
        raise Exception("未找到员工信息")
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    pdf_name = f"新员工座位名牌打印-{timestamp}.pdf"
    pdf_path = os.path.join(output_folder, pdf_name)
    
    c = canvas.Canvas(pdf_path, pagesize=portrait(A4))
    page_width, page_height = portrait(A4)
    badge_height = 278
    
    for i, emp in enumerate(employees):
        y_top = page_height - (i % 3) * badge_height
        y_bottom = y_top - badge_height
        
        c.drawImage(ImageReader(bot_img), 0, y_bottom, width=page_width, height=110)
        c.drawImage(ImageReader(top_img), 0, y_top - 55, width=page_width, height=55)
        
        photo_file = emp["照片文件命名"]
        photo_path = os.path.join(photo_folder, photo_file) if photo_file else ""
        if not photo_file or not os.path.exists(photo_path):
            if os.path.exists(default_photo):
                photo_path = default_photo
            else:
                continue
        
        c.drawImage(ImageReader(photo_path), 27, y_top - 205, width=125, height=125)
        
        c.setLineWidth(2)
        c.line(305, y_top - 95, 305, y_top - 200)
        
        c.setFont("Chinese-Bold", 38)
        name = emp["姓名"] or ""
        name_x = 145 + (150 - c.stringWidth(name, "Chinese-Bold", 38)) / 2
        c.drawString(name_x, y_top - 150, name)
        
        c.setFont("Chinese-Bold", 14)
        pinyin_text = emp["拼音"] or ""
        if pinyin_text:
            pinyin_x = 145 + (150 - c.stringWidth(pinyin_text, "Chinese-Bold", 14)) / 2
            c.drawString(pinyin_x, y_top - 180, pinyin_text)
        
        c.setFont("Chinese", 17)
        c.drawString(340, y_top - 125, f"工号: {emp['工号'] or ''}")
        c.drawString(340, y_top - 155, f"部门: {emp['部门'] or ''}")
        c.drawString(340, y_top - 185, f"岗位: {emp['岗位'] or ''}")
        
        if (i + 1) % 3 == 0 and i != len(employees) - 1:
            c.showPage()
    
    c.save()
    return pdf_path


def generate_suzhou_seat_badge(excel_path, photo_folder, output_folder, assets_folder, fonts_folder, department_list=None):
    """
    生成苏州座位名牌的主函数（可被 Web 后端调用）
    
    参数:
        excel_path: Excel文件路径（包含员工信息）
        photo_folder: 员工照片文件夹路径
        output_folder: 输出文件夹路径
        assets_folder: 资源文件夹路径（包含 top.jpg, bot.jpg, bird.jpg）
        fonts_folder: 字体文件夹路径（包含 msyh.ttc, msyhbd.ttc）
        department_list: 预设部门列表（可选）
    
    返回:
        生成的 PDF 文件路径
    """
    # 在内存中处理数据，不生成中间文件
    employees = process_data_in_memory(excel_path, department_list)
    
    # 直接从员工列表生成PDF
    pdf_path = generate_pdf_from_employees(employees, photo_folder, assets_folder, output_folder, fonts_folder)
    
    return pdf_path


# ========== 以下为保留原有功能的代码（兼容 GUI） ==========

def process_data_copy(data_folder, output_folder, department_list=None):
    """处理原始Excel数据，生成BaseInformation.xlsx（保留原有功能）"""
    source_file = find_excel_file(data_folder)
    
    source_df = pd.read_excel(source_file)
    required_cols = ["姓名", "预入职工号", "岗位", "组织全路径"]
    for col in required_cols:
        if col not in source_df.columns:
            raise Exception(f"原始Excel缺少列: {col}")
    
    result_data = []
    for _, row in source_df.iterrows():
        name = row["姓名"]
        emp_id = row["预入职工号"]
        position = row["岗位"]
        org_path = row["组织全路径"]
        
        department = extract_department(org_path, department_list)
        
        result_data.append({
            "姓名": name,
            "工号": emp_id,
            "岗位": position,
            "部门": department,
            "拼音": "",
            "照片文件命名": ""
        })
    
    result_df = pd.DataFrame(
        result_data,
        columns=["姓名", "工号", "岗位", "部门", "拼音", "照片文件命名"]
    )
    
    target_file = os.path.join(output_folder, "BaseInformation.xlsx")
    with pd.ExcelWriter(target_file, engine='openpyxl', mode='w') as writer:
        result_df.to_excel(writer, index=False)
    
    return target_file


def process_excel(base_info_path):
    """处理BaseInformation.xlsx，生成拼音和照片文件名（保留原有功能）"""
    if not os.path.exists(base_info_path):
        raise Exception(f"未找到文件: {base_info_path}")
    
    wb = load_workbook(base_info_path)
    sheet = wb.active
    count = 0
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        name = row[0].value
        emp_id = row[1].value
        
        pinyin_text = ""
        if name:
            surname = name[0]
            given_name = name[1:]
            surname_pinyin = pinyin(surname, style=Style.NORMAL)[0][0].capitalize()
            given_pinyin = ""
            if given_name:
                given_parts = pinyin(given_name, style=Style.NORMAL)
                given_pinyin = given_parts[0][0].capitalize() + ''.join([p[0] for p in given_parts[1:]])
            pinyin_text = f"{surname_pinyin} {given_pinyin}".strip()
        
        photo_name = f"{emp_id}.jpg" if emp_id else ""
        
        sheet.cell(row=row[0].row, column=5, value=pinyin_text)
        sheet.cell(row=row[0].row, column=6, value=photo_name)
        
        count += 1
    
    wb.save(base_info_path)
    return True


def generate_pdf_from_data(base_info_path, photo_folder, assets_folder, output_folder, fonts_folder):
    """生成最终的PDF名牌文件（保留原有功能）"""
    setup_pdf_fonts(fonts_folder)
    
    top_img = os.path.join(assets_folder, "top.jpg")
    bot_img = os.path.join(assets_folder, "bot.jpg")
    default_photo = os.path.join(assets_folder, "bird.jpg")
    
    if not os.path.exists(top_img):
        raise Exception(f"未找到top.jpg")
    if not os.path.exists(bot_img):
        raise Exception(f"未找到bot.jpg")
    if not os.path.exists(base_info_path):
        raise Exception(f"未找到BaseInformation.xlsx")
    
    df = pd.read_excel(base_info_path)
    required_cols = ["姓名", "工号", "岗位", "部门", "拼音", "照片文件命名"]
    for col in required_cols:
        if col not in df.columns:
            raise Exception(f"BaseInformation.xlsx缺少列: {col}")
    
    employees = df.to_dict("records")
    if not employees:
        raise Exception("未找到员工信息")
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    pdf_name = f"新员工座位名牌打印-{timestamp}.pdf"
    pdf_path = os.path.join(output_folder, pdf_name)
    
    c = canvas.Canvas(pdf_path, pagesize=portrait(A4))
    page_width, page_height = portrait(A4)
    badge_height = 278
    
    for i, emp in enumerate(employees):
        y_top = page_height - (i % 3) * badge_height
        y_bottom = y_top - badge_height
        
        c.drawImage(ImageReader(bot_img), 0, y_bottom, width=page_width, height=110)
        c.drawImage(ImageReader(top_img), 0, y_top - 55, width=page_width, height=55)
        
        photo_file = emp["照片文件命名"]
        photo_path = os.path.join(photo_folder, photo_file) if photo_file else ""
        if not photo_file or not os.path.exists(photo_path):
            if os.path.exists(default_photo):
                photo_path = default_photo
            else:
                continue
        
        c.drawImage(ImageReader(photo_path), 27, y_top - 205, width=125, height=125)
        
        c.setLineWidth(2)
        c.line(305, y_top - 95, 305, y_top - 200)
        
        c.setFont("Chinese-Bold", 38)
        name = emp["姓名"] or ""
        name_x = 145 + (150 - c.stringWidth(name, "Chinese-Bold", 38)) / 2
        c.drawString(name_x, y_top - 150, name)
        
        c.setFont("Chinese-Bold", 14)
        pinyin_text = emp["拼音"] or ""
        if pinyin_text:
            pinyin_x = 145 + (150 - c.stringWidth(pinyin_text, "Chinese-Bold", 14)) / 2
            c.drawString(pinyin_x, y_top - 180, pinyin_text)
        
        c.setFont("Chinese", 17)
        c.drawString(340, y_top - 125, f"工号: {emp['工号'] or ''}")
        c.drawString(340, y_top - 155, f"部门: {emp['部门'] or ''}")
        c.drawString(340, y_top - 185, f"岗位: {emp['岗位'] or ''}")
        
        if (i + 1) % 3 == 0 and i != len(employees) - 1:
            c.showPage()
    
    c.save()
    return pdf_path


# ========== GUI 界面类（保持原有功能） ==========

class NameBadgeGenerator:
    """新员工座位名牌生成工具类（GUI 界面）"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("新员工座位名牌生成工具")
        self.root.geometry("800x500")
        
        self.create_widgets()
        
        if getattr(sys, 'frozen', False):
            self.script_dir = os.path.dirname(os.path.abspath(sys.executable))
        else:
            self.script_dir = os.path.dirname(os.path.abspath(__file__))
        
        self.assets_folder = os.path.join(self.script_dir, "assets")
        
        project_root = os.path.dirname(os.path.dirname(self.script_dir))
        self.fonts_folder = os.path.join(project_root, "shared", "fonts")
        
        self.setup_fonts()
        
        self.data_folder = ""
        self.department_list = [
            "流程数据与IT部", "人力资源部", "电梯产品事业部", "研发管理部", 
            "集成供应链管理部", "总裁办公室A", "总裁办公室B", "黄秘办公室", 
            "审计部", "知识产权与法务中心", "汇川书院", "财经管理部", 
            "质量管理部", "公司变革项目群", "总部派驻中心", "产品竞争力中心", 
            "数字化事业部", "技术服务中心", "战略与投资发展部", "全球工业自动化BG"
        ]
        
        self.log("欢迎使用新员工座位名牌生成工具！")
        self.log(f"程序所在目录：{self.script_dir}")
        self.log("请将员工照片和Excel表格放在同一个文件夹中使用")
    
    def setup_fonts(self):
        """配置PDF中文字体（优先使用打包的字体）"""
        try:
            if getattr(sys, 'frozen', False):
                resource_dir = sys._MEIPASS
            else:
                resource_dir = self.script_dir
            
            font_file = "msyh.ttc"
            font_bold_file = "msyhbd.ttc"
            
            font_path = os.path.join(resource_dir, font_file)
            font_bold_path = os.path.join(resource_dir, font_bold_file)
            
            if not os.path.exists(font_path):
                font_path = os.path.join(self.fonts_folder, font_file)
            if not os.path.exists(font_bold_path):
                font_bold_path = os.path.join(self.fonts_folder, font_bold_file)
            
            if not os.path.exists(font_path):
                sys_font_path = os.path.join("C:", "Windows", "Fonts")
                font_path = os.path.join(sys_font_path, font_file)
            if not os.path.exists(font_bold_path):
                sys_font_path = os.path.join("C:", "Windows", "Fonts")
                font_bold_path = os.path.join(sys_font_path, font_bold_file)
            
            pdfmetrics.registerFont(TTFont("Chinese", font_path))
            pdfmetrics.registerFont(TTFont("Chinese-Bold", font_bold_path))
            self.log("字体注册成功")
            
        except Exception as e:
            self.log(f"字体注册警告: {str(e)}，可能导致PDF中文显示异常")
    
    def create_widgets(self):
        """创建GUI界面（先于setup_fonts执行，确保log_text存在）"""
        top_frame = ttk.Frame(self.root, padding="10")
        top_frame.pack(fill=tk.X)
        
        self.folder_btn = ttk.Button(
            top_frame, 
            text="选择数据文件夹", 
            command=self.select_data_folder
        )
        self.folder_btn.pack(side=tk.LEFT, padx=5)
        
        self.folder_path_var = tk.StringVar()
        self.folder_path_label = ttk.Label(
            top_frame, 
            textvariable=self.folder_path_var, 
            width=60, 
            relief=tk.SUNKEN
        )
        self.folder_path_label.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        mid_frame = ttk.Frame(self.root, padding="10")
        mid_frame.pack(fill=tk.X)
        
        self.generate_btn = ttk.Button(
            mid_frame, 
            text="生成PDF文件", 
            command=self.generate_final_pdf,
            state=tk.DISABLED
        )
        self.generate_btn.pack(side=tk.LEFT, padx=5)
        
        log_frame = ttk.LabelFrame(self.root, text="操作日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)
    
    def log(self, message):
        """在日志区域添加消息（确保log_text已初始化）"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()
    
    def select_data_folder(self):
        """选择数据文件夹（包含员工照片和Excel表格）"""
        folder = filedialog.askdirectory(title="选择包含员工照片和信息表格的文件夹")
        if folder:
            self.data_folder = folder
            self.folder_path_var.set(folder)
            self.log(f"已选择数据文件夹: {folder}")
            
            try:
                excel_files = [f for f in os.listdir(folder) 
                              if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
                image_files = [f for f in os.listdir(folder) 
                              if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
                
                if len(excel_files) == 0:
                    messagebox.warning("警告", "未找到Excel文件！")
                elif len(excel_files) > 1:
                    messagebox.warning("警告", f"找到{len(excel_files)}个Excel文件，需且只能有1个！")
                else:
                    self.log(f"找到Excel文件: {excel_files[0]}")
                    self.log(f"找到{len(image_files)}个图片文件")
                    self.generate_btn.config(state=tk.NORMAL)
            except Exception as e:
                self.log(f"检查文件夹失败: {str(e)}")
                messagebox.showerror("错误", f"检查文件夹时出错: {str(e)}")
    
    def find_excel_file(self, folder):
        """查找文件夹中的Excel文件（确保只有一个）"""
        return find_excel_file(folder)
    
    def extract_department(self, org_path):
        """从组织全路径中提取部门信息"""
        return extract_department(org_path, self.department_list)
    
    def process_data_copy(self):
        """处理原始Excel数据，生成BaseInformation.xlsx"""
        source_file = self.find_excel_file(self.data_folder)
        self.log(f"读取原始数据: {os.path.basename(source_file)}")
        
        source_df = pd.read_excel(source_file)
        required_cols = ["姓名", "预入职工号", "岗位", "组织全路径"]
        for col in required_cols:
            if col not in source_df.columns:
                raise Exception(f"原始Excel缺少列: {col}")
        
        result_data = []
        for _, row in source_df.iterrows():
            name = row["姓名"]
            emp_id = row["预入职工号"]
            position = row["岗位"]
            org_path = row["组织全路径"]
            
            department = self.extract_department(org_path)
            if department not in self.department_list:
                self.log(f"警告: 员工“{name}”的部门可能不正确（{department}）")
            
            result_data.append({
                "姓名": name,
                "工号": emp_id,
                "岗位": position,
                "部门": department,
                "拼音": "",
                "照片文件命名": ""
            })
        
        result_df = pd.DataFrame(
            result_data,
            columns=["姓名", "工号", "岗位", "部门", "拼音", "照片文件命名"]
        )
        
        target_file = os.path.join(self.script_dir, "BaseInformation.xlsx")
        with pd.ExcelWriter(target_file, engine='openpyxl', mode='w') as writer:
            result_df.to_excel(writer, index=False)
        
        self.log(f"已生成BaseInformation.xlsx（{len(result_data)}条记录）")
        return target_file
    
    def process_excel(self, base_info_path):
        """处理BaseInformation.xlsx，生成拼音和照片文件名"""
        return process_excel(base_info_path)
    
    def generate_pdf(self):
        """生成最终的PDF名牌文件"""
        try:
            if getattr(sys, 'frozen', False):
                resource_dir = sys._MEIPASS
            else:
                resource_dir = self.script_dir
            
            base_info_path = os.path.join(self.script_dir, "BaseInformation.xlsx")
            photo_folder = self.data_folder
            top_img = os.path.join(self.assets_folder, "top.jpg")
            bot_img = os.path.join(self.assets_folder, "bot.jpg")
            default_photo = os.path.join(self.assets_folder, "bird.jpg")
            
            if not os.path.exists(top_img):
                raise Exception(f"未找到top.jpg，请确保已打包")
            if not os.path.exists(bot_img):
                raise Exception(f"未找到bot.jpg，请确保已打包")
            if not os.path.exists(base_info_path):
                raise Exception(f"未找到BaseInformation.xlsx")
            
            df = pd.read_excel(base_info_path)
            required_cols = ["姓名", "工号", "岗位", "部门", "拼音", "照片文件命名"]
            for col in required_cols:
                if col not in df.columns:
                    raise Exception(f"BaseInformation.xlsx缺少列: {col}")
            
            employees = df.to_dict("records")
            if not employees:
                raise Exception("未找到员工信息")
            
            self.log(f"开始生成PDF（共{len(employees)}名员工）")
            
            timestamp = datetime.now().strftime("%Y%m%d%H%M")
            pdf_name = f"新员工座位名牌打印-{timestamp}.pdf"
            pdf_path = os.path.join(self.script_dir, pdf_name)
            
            c = canvas.Canvas(pdf_path, pagesize=portrait(A4))
            page_width, page_height = portrait(A4)
            badge_height = 278
            
            for i, emp in enumerate(employees):
                y_top = page_height - (i % 3) * badge_height
                y_bottom = y_top - badge_height
                
                c.drawImage(ImageReader(bot_img), 0, y_bottom, width=page_width, height=110)
                c.drawImage(ImageReader(top_img), 0, y_top - 55, width=page_width, height=55)
                
                photo_file = emp["照片文件命名"]
                photo_path = os.path.join(photo_folder, photo_file) if photo_file else ""
                if not photo_file or not os.path.exists(photo_path):
                    if os.path.exists(default_photo):
                        photo_path = default_photo
                        self.log(f"警告: 员工“{emp['姓名']}”的照片缺失，使用默认照片")
                    else:
                        self.log(f"警告: 员工“{emp['姓名']}”的照片缺失")
                        continue
                
                c.drawImage(ImageReader(photo_path), 27, y_top - 205, width=125, height=125)
                
                c.setLineWidth(2)
                c.line(305, y_top - 95, 305, y_top - 200)
                
                c.setFont("Chinese-Bold", 38)
                name = emp["姓名"] or ""
                name_x = 145 + (150 - c.stringWidth(name, "Chinese-Bold", 38)) / 2
                c.drawString(name_x, y_top - 150, name)
                
                c.setFont("Chinese-Bold", 14)
                pinyin_text = emp["拼音"] or ""
                if pinyin_text:
                    pinyin_x = 145 + (150 - c.stringWidth(pinyin_text, "Chinese-Bold", 14)) / 2
                    c.drawString(pinyin_x, y_top - 180, pinyin_text)
                
                c.setFont("Chinese", 17)
                c.drawString(340, y_top - 125, f"工号: {emp['工号'] or ''}")
                c.drawString(340, y_top - 155, f"部门: {emp['部门'] or ''}")
                c.drawString(340, y_top - 185, f"岗位: {emp['岗位'] or ''}")
                
                if (i + 1) % 3 == 0 and i != len(employees) - 1:
                    c.showPage()
            
            c.save()
            self.log(f"PDF生成成功！路径：{pdf_path}")
            return pdf_path
            
        except Exception as e:
            self.log(f"PDF生成失败: {str(e)}")
            raise Exception(f"PDF生成失败: {str(e)}")
    
    def generate_final_pdf(self):
        """生成PDF的主流程（串联所有步骤）"""
        try:
            if not self.data_folder:
                messagebox.showwarning("提示", "请先选择数据文件夹")
                return
            
            self.log("===== 开始生成流程 =====")
            base_info_path = self.process_data_copy()
            self.process_excel(base_info_path)
            pdf_path = self.generate_pdf()
            self.log("===== 生成完成 =====")
            messagebox.showinfo("成功", f"PDF已生成：\n{pdf_path}")
            
            try:
                os.startfile(pdf_path)
            except:
                self.log("无法自动打开PDF，请手动查看")
            
        except Exception as e:
            self.log(f"操作失败: {str(e)}")
            messagebox.showerror("失败", f"生成过程出错：\n{str(e)}")


# 程序入口（当直接运行该脚本时执行）
if __name__ == "__main__":
    root = tk.Tk()
    app = NameBadgeGenerator(root)
    root.mainloop()
