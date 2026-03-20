# 导入操作系统相关功能模块（文件/目录操作等）
import os
# 导入系统相关功能模块（命令行参数、路径等）
import sys
# 导入与C语言兼容的类型模块（可能用于系统级操作）
import ctypes
# 导入tkinter库，用于创建GUI界面
import tkinter as tk
# 从tkinter导入文件对话框、消息框、带滚动条的文本框组件
from tkinter import filedialog, messagebox, scrolledtext
# 从tkinter导入ttk模块（提供更美观的界面组件）
from tkinter import ttk
# 导入pandas库，用于处理Excel表格数据
import pandas as pd
# 从openpyxl导入load_workbook，用于读取和写入Excel文件
from openpyxl import load_workbook
# 从pypinyin导入拼音转换相关功能
from pypinyin import pinyin, Style
# 从reportlab导入pdfgen模块，用于生成PDF文件
from reportlab.pdfgen import canvas
# 从reportlab导入页面大小相关常量（A4纸、纵向等）
from reportlab.lib.pagesizes import A4, portrait
# 从reportlab导入图片处理工具
from reportlab.lib.utils import ImageReader
# 从reportlab导入PDF字体相关模块
from reportlab.pdfbase import pdfmetrics
# 从reportlab导入TrueType字体处理类
from reportlab.pdfbase.ttfonts import TTFont
# 从datetime导入datetime类，用于处理时间
from datetime import datetime
# 导入idlelib（可能用于备用的界面支持，此处未直接使用）
import idlelib


# 定义新员工座位名牌生成工具类
class NameBadgeGenerator:
    # 类的初始化方法，创建GUI窗口时调用
    def __init__(self, root):
        # 保存主窗口对象
        self.root = root
        # 设置窗口标题
        self.root.title("新员工座位名牌生成工具")
        # 设置窗口初始大小（宽度800像素，高度500像素）
        self.root.geometry("800x500")
        
        # 先初始化界面组件（确保日志文本框先存在，后续日志输出才不会出错）
        self.create_widgets()
        
        # 确定程序所在目录（区分打包后的程序和未打包的脚本）
        # 如果是打包后的可执行文件（如.exe）
        if getattr(sys, 'frozen', False):
            self.script_dir = os.path.dirname(os.path.abspath(sys.executable))
        # 如果是未打包的Python脚本
        else:
            self.script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 配置PDF中使用的字体（此时日志方法已可用，可输出配置信息）
        self.setup_fonts()
        
        # 初始化数据文件夹路径（未选择时为空）
        self.data_folder = ""
        # 预设部门列表（用于校验提取的部门是否合理）
        self.department_list = [
            "流程数据与IT部", "人力资源部", "电梯产品事业部", "研发管理部", 
            "集成供应链管理部", "总裁办公室A", "总裁办公室B", "黄秘办公室", 
            "审计部", "知识产权与法务中心", "汇川书院", "财经管理部", 
            "质量管理部", "公司变革项目群", "总部派驻中心", "产品竞争力中心", 
            "数字化事业部", "技术服务中心", "战略与投资发展部", "全球工业自动化BG"
        ]
        
        # 输出欢迎日志
        self.log("欢迎使用新员工座位名牌生成工具！")
        # 输出程序所在目录日志
        self.log(f"程序所在目录：{self.script_dir}")
        # 输出使用提示日志
        self.log("请将员工照片和Excel表格放在同一个文件夹中使用")

    # 配置PDF中使用的中文字体（确保中文能正常显示）
    def setup_fonts(self):
        """配置PDF中文字体（优先使用打包的字体）"""
        try:
            # 确定资源文件目录（打包后为临时目录，未打包为脚本目录）
            if getattr(sys, 'frozen', False):
                resource_dir = sys._MEIPASS  # 打包后的资源目录
            else:
                resource_dir = self.script_dir  # 未打包时的脚本目录
            
            # 字体文件名称（需与打包的文件一致）
            font_file = "msyh.ttc"  # 微软雅黑常规字体
            font_bold_file = "msyhbd.ttc"  # 微软雅黑粗体字体
            
            # 优先使用打包的字体文件（即使系统没有也能运行）
            font_path = os.path.join(resource_dir, font_file)
            font_bold_path = os.path.join(resource_dir, font_bold_file)
            
            # 若打包的字体不存在，再尝试从系统字体目录查找
            if not os.path.exists(font_path):
                sys_font_path = os.path.join("C:", "Windows", "Fonts")  # Windows系统字体目录
                font_path = os.path.join(sys_font_path, font_file)
            if not os.path.exists(font_bold_path):
                sys_font_path = os.path.join("C:", "Windows", "Fonts")
                font_bold_path = os.path.join(sys_font_path, font_bold_file)
            
            # 注册字体（供PDF生成时使用）
            pdfmetrics.registerFont(TTFont("Chinese", font_path))
            pdfmetrics.registerFont(TTFont("Chinese-Bold", font_bold_path))
            self.log("字体注册成功")
            
        # 捕获字体配置异常（如字体文件缺失）
        except Exception as e:
            self.log(f"字体注册警告: {str(e)}，可能导致PDF中文显示异常")

    # 创建GUI界面组件
    def create_widgets(self):
        """创建GUI界面（先于setup_fonts执行，确保log_text存在）"""
        # 顶部：选择文件夹区域（使用ttk.Frame作为容器）
        top_frame = ttk.Frame(self.root, padding="10")  # padding设置内边距
        top_frame.pack(fill=tk.X)  # 水平方向填充父容器
        
        # 创建"选择数据文件夹"按钮，点击触发select_data_folder方法
        self.folder_btn = ttk.Button(
            top_frame, 
            text="选择数据文件夹", 
            command=self.select_data_folder
        )
        self.folder_btn.pack(side=tk.LEFT, padx=5)  # 左对齐，水平间距5像素
        
        # 创建用于显示文件夹路径的变量
        self.folder_path_var = tk.StringVar()
        # 创建显示文件夹路径的标签（凹陷样式，类似输入框）
        self.folder_path_label = ttk.Label(
            top_frame, 
            textvariable=self.folder_path_var,  # 绑定路径变量
            width=60,  # 宽度60字符
            relief=tk.SUNKEN  # 凹陷样式
        )
        self.folder_path_label.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)  # 填充剩余水平空间
        
        # 中间：操作按钮区域
        mid_frame = ttk.Frame(self.root, padding="10")
        mid_frame.pack(fill=tk.X)
        
        # 创建"生成PDF文件"按钮，点击触发generate_final_pdf方法，初始状态为禁用
        self.generate_btn = ttk.Button(
            mid_frame, 
            text="生成PDF文件", 
            command=self.generate_final_pdf,
            state=tk.DISABLED  # 初始禁用，选择有效文件夹后启用
        )
        self.generate_btn.pack(side=tk.LEFT, padx=5)
        
        # 底部：日志显示区域（使用带标签的框架）
        log_frame = ttk.LabelFrame(self.root, text="操作日志", padding="10")  # 带"操作日志"标题的框架
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)  # 填充整个剩余空间
        
        # 创建带滚动条的文本框，用于显示日志
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD)  # 自动换行
        self.log_text.pack(fill=tk.BOTH, expand=True)  # 填充整个日志框架
        self.log_text.config(state=tk.DISABLED)  # 初始设置为只读（防止手动编辑）

    # 在日志区域添加消息
    def log(self, message):
        """在日志区域添加消息（确保log_text已初始化）"""
        self.log_text.config(state=tk.NORMAL)  # 临时开启可编辑状态
        # 插入带时间戳的消息（当前时间+消息内容）
        self.log_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)  # 自动滚动到最后一行
        self.log_text.config(state=tk.DISABLED)  # 恢复只读状态
        self.root.update_idletasks()  # 强制刷新界面，立即显示新日志

    # 选择数据文件夹（包含员工照片和Excel表格）
    def select_data_folder(self):
        # 打开文件夹选择对话框，返回选择的文件夹路径
        folder = filedialog.askdirectory(title="选择包含员工照片和信息表格的文件夹")
        if folder:  # 如果用户选择了文件夹（未取消）
            self.data_folder = folder  # 保存文件夹路径
            self.folder_path_var.set(folder)  # 在标签中显示文件夹路径
            self.log(f"已选择数据文件夹: {folder}")  # 记录日志
            
            try:
                # 查找文件夹中所有Excel文件（排除临时文件~$开头的）
                excel_files = [f for f in os.listdir(folder) 
                              if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
                # 查找文件夹中所有图片文件（.jpg/.jpeg/.png，不区分大小写）
                image_files = [f for f in os.listdir(folder) 
                              if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
                
                # 检查Excel文件数量是否合法
                if len(excel_files) == 0:
                    messagebox.warning("警告", "未找到Excel文件！")  # 提示未找到Excel
                elif len(excel_files) > 1:
                    # 提示找到多个Excel（要求只能有一个）
                    messagebox.warning("警告", f"找到{len(excel_files)}个Excel文件，需且只能有1个！")
                else:
                    # 找到单个Excel，记录日志并启用生成按钮
                    self.log(f"找到Excel文件: {excel_files[0]}")
                    self.log(f"找到{len(image_files)}个图片文件")
                    self.generate_btn.config(state=tk.NORMAL)  # 启用生成按钮
            # 捕获文件夹检查过程中的异常
            except Exception as e:
                self.log(f"检查文件夹失败: {str(e)}")
                messagebox.showerror("错误", f"检查文件夹时出错: {str(e)}")

    # 查找文件夹中的Excel文件（确保只有一个）
    def find_excel_file(self, folder):
        # 获取文件夹中所有有效的Excel文件路径
        excel_files = [os.path.join(folder, f) for f in os.listdir(folder)
                      if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
        
        # 如果Excel文件数量不是1，抛出异常
        if len(excel_files) != 1:
            raise Exception(f"找到{len(excel_files)}个Excel文件，需且只能有1个")
        
        return excel_files[0]  # 返回找到的唯一Excel文件路径

    # 从组织全路径中提取部门信息
    def extract_department(self, org_path):
        # 如果组织路径不是字符串（如空值），返回空
        if not isinstance(org_path, str):
            return ""
        
        # 处理以"汇川集团"开头的路径
        if org_path.startswith("汇川集团"):
            parts = org_path.split("/")  # 按"/"分割路径
            if "总部" in parts:  # 如果路径中包含"总部"
                总部索引 = parts.index("总部")  # 找到"总部"的位置
                # 如果总部后还有内容，取总部后的第一个部分作为部门
                if 总部索引 + 1 < len(parts):
                    return parts[总部索引 + 1]
        # 处理其他格式的路径
        else:
            if "/" in org_path:  # 如果包含"/"，取第一个部分作为部门
                return org_path.split("/")[0]
        
        # 无法提取时，直接返回原始路径
        return org_path

    # 处理原始Excel数据，生成BaseInformation.xlsx
    def process_data_copy(self):
        try:
            # 找到数据文件夹中的Excel文件
            source_file = self.find_excel_file(self.data_folder)
            self.log(f"读取原始数据: {os.path.basename(source_file)}")  # 记录日志
            
            # 读取原始Excel文件内容
            source_df = pd.read_excel(source_file)
            # 定义必须包含的列（缺少则无法生成名牌）
            required_cols = ["姓名", "预入职工号", "岗位", "组织全路径"]
            # 检查是否缺少必要的列
            for col in required_cols:
                if col not in source_df.columns:
                    raise Exception(f"原始Excel缺少列: {col}")
            
            # 存储处理后的员工数据
            result_data = []
            # 遍历原始数据中的每一行（每一名员工）
            for _, row in source_df.iterrows():
                name = row["姓名"]  # 姓名
                emp_id = row["预入职工号"]  # 工号
                position = row["岗位"]  # 岗位
                org_path = row["组织全路径"]  # 组织全路径
                
                # 从组织路径中提取部门
                department = self.extract_department(org_path)
                # 如果提取的部门不在预设列表中，记录警告
                if department not in self.department_list:
                    self.log(f"警告: 员工“{name}”的部门可能不正确（{department}）")
                
                # 添加处理后的员工信息到结果列表
                result_data.append({
                    "姓名": name,
                    "工号": emp_id,
                    "岗位": position,
                    "部门": department,
                    "拼音": "",  # 拼音后续处理
                    "照片文件命名": ""  # 照片文件名后续处理
                })
            
            # 将结果列表转换为DataFrame（表格数据）
            result_df = pd.DataFrame(
                result_data,
                columns=["姓名", "工号", "岗位", "部门", "拼音", "照片文件命名"]
            )
            
            # 定义目标文件路径（程序目录下的BaseInformation.xlsx）
            target_file = os.path.join(self.script_dir, "BaseInformation.xlsx")
            # 将处理后的数据写入Excel文件
            with pd.ExcelWriter(target_file, engine='openpyxl', mode='w') as writer:
                result_df.to_excel(writer, index=False)  # 不写入索引列
            
            self.log(f"已生成BaseInformation.xlsx（{len(result_data)}条记录）")
            return target_file  # 返回生成的Excel路径
            
        # 捕获数据处理异常
        except Exception as e:
            self.log(f"数据提取失败: {str(e)}")
            raise Exception(f"数据提取失败: {str(e)}")

    # 处理BaseInformation.xlsx，生成拼音和照片文件名
    def process_excel(self, base_info_path):
        try:
            # 检查文件是否存在
            if not os.path.exists(base_info_path):
                raise Exception(f"未找到文件: {base_info_path}")
            
            # 加载Excel文件
            wb = load_workbook(base_info_path)
            sheet = wb.active  # 获取活动工作表
            count = 0  # 记录处理的行数
            
            # 遍历表格中的行（从第2行开始，跳过表头）
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                name = row[0].value  # 第1列：姓名
                emp_id = row[1].value  # 第2列：工号
                
                pinyin_text = ""  # 初始化拼音
                if name:  # 如果姓名不为空
                    surname = name[0]  # 姓氏（姓名的第一个字符）
                    given_name = name[1:]  # 名字（姓名剩余部分）
                    # 转换姓氏为拼音（首字母大写）
                    surname_pinyin = pinyin(surname, style=Style.NORMAL)[0][0].capitalize()
                    given_pinyin = ""  # 初始化名字拼音
                    if given_name:  # 如果名字不为空
                        # 转换名字为拼音
                        given_parts = pinyin(given_name, style=Style.NORMAL)
                        # 名字拼音首字母大写，后续字母小写（如"晓华"→"Xiaohua"）
                        given_pinyin = given_parts[0][0].capitalize() + ''.join([p[0] for p in given_parts[1:]])
                    # 组合姓氏和名字拼音（如"张三"→"Zhang San"）
                    pinyin_text = f"{surname_pinyin} {given_pinyin}".strip()
                
                # 照片文件命名规则：工号.jpg
                photo_name = f"{emp_id}.jpg" if emp_id else ""
                
                # 将拼音写入第5列
                sheet.cell(row=row[0].row, column=5, value=pinyin_text)
                # 将照片文件名写入第6列
                sheet.cell(row=row[0].row, column=6, value=photo_name)
                
                count += 1  # 计数+1
            
            # 保存修改后的Excel文件
            wb.save(base_info_path)
            self.log(f"已处理{count}条员工信息")
            return True  # 处理成功
            
        # 捕获Excel处理异常
        except Exception as e:
            self.log(f"Excel处理失败: {str(e)}")
            raise Exception(f"Excel处理失败: {str(e)}")

    # 生成最终的PDF名牌文件
    def generate_pdf(self):
        try:
            # 确定资源文件目录（同字体配置逻辑）
            if getattr(sys, 'frozen', False):
                resource_dir = sys._MEIPASS
            else:
                resource_dir = self.script_dir
            
            # 定义所需文件路径
            base_info_path = os.path.join(self.script_dir, "BaseInformation.xlsx")  # 员工信息表
            photo_folder = self.data_folder  # 照片所在文件夹
            top_img = os.path.join(resource_dir, "top.jpg")  # 名牌顶部图片
            bot_img = os.path.join(resource_dir, "bot.jpg")  # 名牌底部图片
            default_photo = os.path.join(resource_dir, "bird.jpg")  # 缺省照片（无照片时使用）
            
            # 检查必要文件是否存在
            if not os.path.exists(top_img):
                raise Exception(f"未找到top.jpg，请确保已打包")
            if not os.path.exists(bot_img):
                raise Exception(f"未找到bot.jpg，请确保已打包")
            if not os.path.exists(base_info_path):
                raise Exception(f"未找到BaseInformation.xlsx")
            
            # 读取员工信息表
            df = pd.read_excel(base_info_path)
            # 检查必要的列是否存在
            required_cols = ["姓名", "工号", "岗位", "部门", "拼音", "照片文件命名"]
            for col in required_cols:
                if col not in df.columns:
                    raise Exception(f"BaseInformation.xlsx缺少列: {col}")
            
            # 将表格数据转换为字典列表（便于遍历）
            employees = df.to_dict("records")
            if not employees:  # 如果没有员工信息
                raise Exception("未找到员工信息")
            
            self.log(f"开始生成PDF（共{len(employees)}名员工）")  # 记录日志
            
            # 生成带时间戳的PDF文件名（避免重名）
            timestamp = datetime.now().strftime("%Y%m%d%H%M")
            pdf_name = f"新员工座位名牌打印-{timestamp}.pdf"
            pdf_path = os.path.join(self.script_dir, pdf_name)  # PDF保存路径
            
            # 创建PDF画布（A4纸，纵向）
            c = canvas.Canvas(pdf_path, pagesize=portrait(A4))
            page_width, page_height = portrait(A4)  # 获取页面宽高
            badge_height = 278  # 每个名牌的高度（A4纵向可放3个）
            
            # 遍历每一名员工，绘制名牌
            for i, emp in enumerate(employees):
                # 计算当前名牌的顶部和底部Y坐标（每3个换一页）
                y_top = page_height - (i % 3) * badge_height
                y_bottom = y_top - badge_height
                
                # 绘制名牌底部图片
                c.drawImage(ImageReader(bot_img), 0, y_bottom, width=page_width, height=110)
                # 绘制名牌顶部图片
                c.drawImage(ImageReader(top_img), 0, y_top - 55, width=page_width, height=55)
                
                # 处理员工照片
                photo_file = emp["照片文件命名"]  # 照片文件名
                # 照片完整路径
                photo_path = os.path.join(photo_folder, photo_file) if photo_file else ""
                # 如果照片不存在，使用默认照片
                if not photo_file or not os.path.exists(photo_path):
                    if os.path.exists(default_photo):
                        photo_path = default_photo
                        self.log(f"警告: 员工“{emp['姓名']}”的照片缺失，使用默认照片")
                    else:
                        self.log(f"警告: 员工“{emp['姓名']}”的照片缺失")
                        continue  # 跳过无照片且无默认照片的员工
                
                # 绘制员工照片（位置：左27，上y_top-205，宽125，高125）
                c.drawImage(ImageReader(photo_path), 27, y_top - 205, width=125, height=125)
                
                # 绘制分隔线（宽度2像素）
                c.setLineWidth(2)
                c.line(305, y_top - 95, 305, y_top - 200)  # 竖线：x=305，y从y_top-95到y_top-200
                
                # 绘制姓名（粗体，38号字）
                c.setFont("Chinese-Bold", 38)
                name = emp["姓名"] or ""  # 姓名（为空时显示空）
                # 计算姓名居中的X坐标
                name_x = 145 + (150 - c.stringWidth(name, "Chinese-Bold", 38)) / 2
                c.drawString(name_x, y_top - 150, name)  # 绘制姓名
                
                # 绘制拼音（粗体，14号字）
                c.setFont("Chinese-Bold", 14)
                pinyin_text = emp["拼音"] or ""
                if pinyin_text:  # 如果拼音不为空
                    # 计算拼音居中的X坐标
                    pinyin_x = 145 + (150 - c.stringWidth(pinyin_text, "Chinese-Bold", 14)) / 2
                    c.drawString(pinyin_x, y_top - 180, pinyin_text)  # 绘制拼音
                
                # 绘制其他信息（常规字体，17号字）
                c.setFont("Chinese", 17)
                c.drawString(340, y_top - 125, f"工号: {emp['工号'] or ''}")  # 工号
                c.drawString(340, y_top - 155, f"部门: {emp['部门'] or ''}")  # 部门
                c.drawString(340, y_top - 185, f"岗位: {emp['岗位'] or ''}")  # 岗位
                
                # 每3个员工换一页（最后一页不额外换页）
                if (i + 1) % 3 == 0 and i != len(employees) - 1:
                    c.showPage()  # 新建一页
            
            # 保存PDF文件
            c.save()
            self.log(f"PDF生成成功！路径：{pdf_path}")
            return pdf_path  # 返回生成的PDF路径
            
        # 捕获PDF生成异常
        except Exception as e:
            self.log(f"PDF生成失败: {str(e)}")
            raise Exception(f"PDF生成失败: {str(e)}")

    # 生成PDF的主流程（串联所有步骤）
    def generate_final_pdf(self):
        try:
            # 如果未选择数据文件夹，提示用户
            if not self.data_folder:
                messagebox.showwarning("提示", "请先选择数据文件夹")
                return
            
            self.log("===== 开始生成流程 =====")  # 记录流程开始
            # 步骤1：处理原始数据，生成BaseInformation.xlsx
            base_info_path = self.process_data_copy()
            # 步骤2：处理BaseInformation.xlsx，补充拼音和照片名
            self.process_excel(base_info_path)
            # 步骤3：生成最终的PDF文件
            pdf_path = self.generate_pdf()
            self.log("===== 生成完成 =====")  # 记录流程结束
            # 提示用户生成成功
            messagebox.showinfo("成功", f"PDF已生成：\n{pdf_path}")
            
            # 尝试自动打开生成的PDF文件
            try:
                os.startfile(pdf_path)
            except:
                self.log("无法自动打开PDF，请手动查看")
            
        # 捕获整个流程中的异常
        except Exception as e:
            self.log(f"操作失败: {str(e)}")
            messagebox.showerror("失败", f"生成过程出错：\n{str(e)}")


# 程序入口（当直接运行该脚本时执行）
if __name__ == "__main__":
    root = tk.Tk()  # 创建Tkinter主窗口
    app = NameBadgeGenerator(root)  # 创建应用实例
    root.mainloop()  # 启动GUI事件循环（保持窗口运行）
