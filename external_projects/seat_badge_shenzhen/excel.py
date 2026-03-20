# -*- coding: utf-8 -*-
import sys
import io
import os
import ctypes
from openpyxl import load_workbook

# 强制设置Windows控制台编码为UTF-8
if sys.platform == 'win32':
    try:
        ctypes.windll.kernel32.SetConsoleCP(65001)
        ctypes.windll.kernel32.SetConsoleOutputCP(65001)
    except:
        pass

# 针对IDLE环境调整stdout编码
if 'idlelib' in sys.modules:  # 判断是否在IDLE中运行
    sys.stdout = io.TextIOWrapper(sys.stdout, encoding='utf-8', errors='ignore')
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='ignore')

# Excel文件路径
file_path = r"E:\AI SpaceX\employNEW-zongbu\BaseInformation.xlsx"

# 检查文件是否存在
if not os.path.exists(file_path):
    print(f"错误：文件不存在 - {file_path}")
    sys.exit(1)

# 加载并修改Excel
wb = load_workbook(file_path)
sheet = wb.active

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    job_id = row[1].value  # 获取B列的工号

    # 生成照片文件命名（格式：工号.jpg）
    photo_name = f"{job_id}.jpg" if job_id else ""

    # 仅写入照片文件命名到E列（第5列）
    sheet.cell(row=row[0].row, column=5, value=photo_name)

wb.save(file_path)
print("处理完成，已更新文件：", file_path)
