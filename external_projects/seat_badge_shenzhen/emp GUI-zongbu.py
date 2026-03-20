import os
import sys
import ctypes
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
import pandas as pd
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime

class NameplateGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("新员工座位名牌生成工具")
        self.root.geometry("800x500")
        
        # 先初始化界面组件（确保日志组件优先创建）
        self.create_widgets()
        
        # 确定程序运行目录（支持exe打包）
        if getattr(sys, 'frozen', False):
            self.program_dir = os.path.dirname(os.path.abspath(sys.executable))
            self.resource_dir = sys._MEIPASS  # 打包后的资源目录
        else:
            self.program_dir = os.path.dirname(os.path.abspath(__file__))
            self.resource_dir = self.program_dir  # 开发时资源目录与程序目录一致
        
        # 初始化配置
        self.data_folder = ""  # 用户选择的包含照片和表格的文件夹
        self.department_list = [
            "流程数据与IT部", "人力资源部", "电梯产品事业部", "研发管理部", 
            "集成供应链管理部", "总裁办公室A", "总裁办公室B", "黄秘办公室", 
            "审计部", "知识产权与法务中心", "汇川书院", "财经管理部", 
            "质量管理部", "公司变革项目群", "总部派驻中心", "产品竞争力中心", 
            "数字化事业部", "技术服务中心", "战略与投资发展部", "全球工业自动化BG"
        ]
        
        # 初始化字体和日志
        self.setup_fonts()
        self.log("欢迎使用新员工座位名牌生成工具！")
        self.log(f"程序运行目录：{self.program_dir}")
        self.log("请将员工照片和Excel表格放在同一个文件夹中，然后点击【选择数据文件夹】")

    def create_widgets(self):
        """创建GUI界面组件"""
        # 顶部：文件夹选择区域
        top_frame = ttk.Frame(self.root, padding="10")
        top_frame.pack(fill=tk.X)
        
        self.select_folder_btn = ttk.Button(
            top_frame, 
            text="选择数据文件夹", 
            command=self.select_data_folder
        )
        self.select_folder_btn.pack(side=tk.LEFT, padx=5)
        
        self.folder_path_var = tk.StringVar()
        self.folder_path_label = ttk.Label(
            top_frame, 
            textvariable=self.folder_path_var, 
            width=60,
            relief=tk.SUNKEN
        )
        self.folder_path_label.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 中间：操作按钮区域
        mid_frame = ttk.Frame(self.root, padding="10")
        mid_frame.pack(fill=tk.X)
        
        self.generate_btn = ttk.Button(
            mid_frame, 
            text="生成PDF文件", 
            command=self.start_generation,
            state=tk.DISABLED  # 初始禁用，选择文件夹后启用
        )
        self.generate_btn.pack(side=tk.LEFT, padx=5)
        
        # 底部：日志显示区域
        log_frame = ttk.LabelFrame(self.root, text="操作日志", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)  # 初始只读

    def log(self, message):
        """在日志区域添加带时间戳的消息"""
        self.log_text.config(state=tk.NORMAL)
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)  # 自动滚动到最新日志
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()  # 实时刷新界面

    def setup_fonts(self):
        """配置PDF中文字体（优先使用资源目录字体，兼容打包后运行）"""
        try:
            # 字体文件名称（需与程序目录下的字体文件一致）
            font_names = {
                "regular": "msyh.ttc",    # 微软雅黑常规
                "bold": "msyhbd.ttc"      # 微软雅黑粗体
            }
            
            # 构建字体路径（优先资源目录，再系统字体目录）
            font_paths = {}
            for font_type, font_file in font_names.items():
                # 检查资源目录
                resource_font = os.path.join(self.resource_dir, font_file)
                if os.path.exists(resource_font):
                    font_paths[font_type] = resource_font
                else:
                    # 检查系统字体目录
                    sys_font = os.path.join("C:", "Windows", "Fonts", font_file)
                    if os.path.exists(sys_font):
                        font_paths[font_type] = sys_font
                    else:
                        raise Exception(f"未找到字体文件：{font_file}")
            
            # 注册字体
            pdfmetrics.registerFont(TTFont("Chinese", font_paths["regular"]))
            pdfmetrics.registerFont(TTFont("Chinese-Bold", font_paths["bold"]))
            self.log("中文字体注册成功，支持PDF中文显示")
            
        except Exception as e:
            self.log(f"字体配置警告：{str(e)}，可能导致PDF中文显示异常")

    def select_data_folder(self):
        """选择包含照片和Excel表格的数据文件夹"""
        folder = filedialog.askdirectory(title="选择数据文件夹（包含照片和Excel表格）")
        if not folder:
            return  # 用户取消选择
        
        self.data_folder = folder
        self.folder_path_var.set(folder)
        self.log(f"已选择数据文件夹：{folder}")
        
        # 验证文件夹内容
        try:
            # 检查Excel文件（仅允许一个有效Excel）
            excel_files = [
                f for f in os.listdir(folder)
                if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')  # 排除临时文件
            ]
            
            # 检查图片文件
            image_files = [
                f for f in os.listdir(folder)
                if f.lower().endswith(('.jpg', '.jpeg', '.png'))
            ]
            
            # 验证Excel数量
            if len(excel_files) == 0:
                messagebox.warning("警告", "所选文件夹中未找到Excel文件！")
                self.generate_btn.config(state=tk.DISABLED)
            elif len(excel_files) > 1:
                messagebox.warning("警告", f"所选文件夹中找到{len(excel_files)}个Excel文件，仅允许一个！")
                self.generate_btn.config(state=tk.DISABLED)
            else:
                self.log(f"找到Excel文件：{excel_files[0]}")
                self.log(f"找到{len(image_files)}个图片文件")
                self.generate_btn.config(state=tk.NORMAL)  # 验证通过，启用生成按钮
                
        except Exception as e:
            self.log(f"文件夹验证失败：{str(e)}")
            messagebox.showerror("错误", f"检查文件夹时出错：{str(e)}")
            self.generate_btn.config(state=tk.DISABLED)

    def find_excel_file(self, folder):
        """从文件夹中获取唯一的Excel文件路径"""
        excel_files = [
            os.path.join(folder, f)
            for f in os.listdir(folder)
            if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')
        ]
        
        if len(excel_files) != 1:
            raise Exception(f"预期1个Excel文件，实际找到{len(excel_files)}个")
        
        return excel_files[0]

    def extract_department(self, org_path):
        """从组织全路径中提取部门信息"""
        if not isinstance(org_path, str):
            return ""
        
        # 处理"汇川集团"开头的路径
        if org_path.startswith("汇川集团"):
            parts = org_path.split("/")
            if "总部" in parts:
                总部索引 = parts.index("总部")
                if 总部索引 + 1 < len(parts):
                    return parts[总部索引 + 1]
        
        # 处理其他路径
        if "/" in org_path:
            return org_path.split("/")[0]
        
        # 默认返回原始路径
        return org_path

    def process_data_copy(self):
        """执行dataCopy.py逻辑：提取数据到BaseInformation.xlsx"""
        try:
            # 1. 找到源Excel文件
            source_excel = self.find_excel_file(self.data_folder)
            self.log(f"读取原始数据：{os.path.basename(source_excel)}")
            
            # 2. 读取并验证数据
            source_df = pd.read_excel(source_excel)
            required_cols = ["姓名", "预入职工号", "岗位", "组织全路径"]
            for col in required_cols:
                if col not in source_df.columns:
                    raise Exception(f"原始Excel缺少必要列：{col}")
            
            # 3. 提取并处理数据
            result_data = []
            for _, row in source_df.iterrows():
                name = row["姓名"]
                emp_id = row["预入职工号"]
                position = row["岗位"]
                org_path = row["组织全路径"]
                
                # 提取部门
                department = self.extract_department(org_path)
                if department not in self.department_list:
                    self.log(f"警告：员工“{name}”的部门信息可能不正确（{department}）")
                
                result_data.append({
                    "姓名": name,
                    "工号": emp_id,
                    "岗位": position,
                    "部门": department,
                    "照片文件命名": ""  # 预留字段，后续由excel.py逻辑填充
                })
            
            # 4. 生成BaseInformation.xlsx
            base_info_path = os.path.join(self.program_dir, "BaseInformation.xlsx")
            result_df = pd.DataFrame(
                result_data,
                columns=["姓名", "工号", "岗位", "部门", "照片文件命名"]
            )
            with pd.ExcelWriter(base_info_path, engine='openpyxl', mode='w') as writer:
                result_df.to_excel(writer, index=False)
            
            self.log(f"已生成BaseInformation.xlsx（{len(result_data)}条记录）")
            return base_info_path
            
        except Exception as e:
            raise Exception(f"数据提取失败：{str(e)}")

    def process_excel(self, base_info_path):
        """执行excel.py逻辑：生成照片文件名"""
        try:
            if not os.path.exists(base_info_path):
                raise Exception(f"文件不存在：{base_info_path}")
            
            # 加载Excel并处理
            wb = load_workbook(base_info_path)
            sheet = wb.active
            processed_count = 0
            
            # 从第2行开始处理（跳过表头）
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                emp_id = row[1].value  # B列：工号
                photo_name = f"{emp_id}.jpg" if emp_id else ""
                sheet.cell(row=row[0].row, column=5, value=photo_name)  # E列：照片文件命名
                processed_count += 1
            
            wb.save(base_info_path)
            self.log(f"已为{processed_count}条记录生成照片文件名")
            return True
            
        except Exception as e:
            raise Exception(f"Excel处理失败：{str(e)}")

    def generate_pdf(self):
        """执行emp.py逻辑：生成PDF名牌"""
        try:
            # 1. 验证必要文件
            base_info_path = os.path.join(self.program_dir, "BaseInformation.xlsx")
            top_img_path = os.path.join(self.resource_dir, "top.jpg")  # 背景图
            default_photo_path = os.path.join(self.resource_dir, "bird.jpg")  # 默认照片
            
            for file_path, desc in [
                (base_info_path, "员工信息表BaseInformation.xlsx"),
                (top_img_path, "背景图top.jpg"),
                (default_photo_path, "默认照片bird.jpg")
            ]:
                if not os.path.exists(file_path):
                    raise Exception(f"缺少必要文件：{desc}（路径：{file_path}）")
            
            # 2. 读取员工信息
            df = pd.read_excel(base_info_path)
            required_cols = ["姓名", "工号", "岗位", "部门", "照片文件命名"]
            for col in required_cols:
                if col not in df.columns:
                    raise Exception(f"BaseInformation.xlsx缺少列：{col}")
            
            employees = df.to_dict("records")
            if not employees:
                raise Exception("BaseInformation.xlsx中未找到员工数据")
            
            self.log(f"开始生成PDF（共{len(employees)}名员工）")
            
            # 3. 初始化PDF
            timestamp = datetime.now().strftime("%Y%m%d%H%M")
            pdf_name = f"新员工座位铭牌打印-{timestamp}.pdf"
            pdf_path = os.path.join(self.program_dir, pdf_name)
            c = canvas.Canvas(pdf_path, pagesize=portrait(A4))
            page_width, page_height = portrait(A4)
            
            # 4. 铭牌布局参数（保留原emp.py的布局设置）
            bg_margin = 35
            bg_top_margin = 30
            bg_height = 260
            bg_width = page_width - 2 * bg_margin
            badge_height = bg_height
            downLengh = 10
            rightLengh = 10
            text_color = 0.2
            
            # 5. 循环生成每个铭牌
            for i, emp in enumerate(employees):
                # 计算当前铭牌位置（每页3个）
                y_top = page_height - bg_top_margin - (i % 3) * badge_height
                y_bottom = y_top - badge_height
                
                # 绘制背景图
                bg_img = ImageReader(top_img_path)
                c.drawImage(bg_img, bg_margin, y_bottom, width=bg_width, height=bg_height)
                
                # 处理员工照片
                photo_file = emp["照片文件命名"]
                photo_path = os.path.join(self.data_folder, photo_file) if photo_file else ""
                if not photo_file or not os.path.exists(photo_path):
                    photo_path = default_photo_path
                    self.log(f"警告：员工“{emp['姓名']}”照片缺失，使用默认照片")
                
                # 绘制照片
                c.drawImage(photo_path, bg_margin + 40, y_bottom + 45, width=150, height=150)
                
                # 绘制姓名（粗体）
                c.setFont("Chinese-Bold", 32)
                c.setFillColorRGB(0, 0, 0)  # 黑色
                name = emp["姓名"] or ""
                name_width = c.stringWidth(name, "Chinese-Bold", 32)
                name_x = bg_margin + 190 + (140 - name_width) / 2  # 居中
                c.drawString(bg_margin + 220 + rightLengh, y_bottom + 160 - downLengh, name)
                
                # 绘制工号
                c.setFont("Chinese-Bold", 17)
                c.setFillColorRGB(text_color, text_color, text_color)
                c.drawString(bg_margin + 220 + rightLengh, y_bottom + 130 - downLengh, "工号: ")
                c.setFont("Chinese", 17)
                c.drawString(
                    bg_margin + 220 + rightLengh + c.stringWidth("工号: ", "Chinese-Bold", 17),
                    y_bottom + 130 - downLengh,
                    str(emp["工号"] or "")
                )
                
                # 绘制部门
                c.setFont("Chinese-Bold", 17)
                c.setFillColorRGB(text_color, text_color, text_color)
                c.drawString(bg_margin + 220 + rightLengh, y_bottom + 100 - downLengh, "部门: ")
                c.setFont("Chinese", 17)
                c.drawString(
                    bg_margin + 220 + rightLengh + c.stringWidth("部门: ", "Chinese-Bold", 17),
                    y_bottom + 100 - downLengh,
                    str(emp["部门"] or "")
                )
                
                # 绘制岗位
                c.setFont("Chinese-Bold", 17)
                c.setFillColorRGB(text_color, text_color, text_color)
                c.drawString(bg_margin + 220 + rightLengh, y_bottom + 70 - downLengh, "岗位: ")
                c.setFont("Chinese", 17)
                c.drawString(
                    bg_margin + 220 + rightLengh + c.stringWidth("岗位: ", "Chinese-Bold", 17),
                    y_bottom + 70 - downLengh,
                    str(emp["岗位"] or "")
                )
                
                # 分页（每3个换一页）
                if (i + 1) % 3 == 0 and i != len(employees) - 1:
                    c.showPage()
            
            # 6. 保存PDF
            c.save()
            self.log(f"PDF生成成功：{pdf_path}")
            return pdf_path
            
        except Exception as e:
            raise Exception(f"PDF生成失败：{str(e)}")

    def start_generation(self):
        """启动完整生成流程（按顺序执行三步操作）"""
        if not self.data_folder:
            messagebox.showwarning("提示", "请先选择数据文件夹")
            return
        
        try:
            self.log("===== 开始生成流程 =====")
            
            # 步骤1：数据提取（dataCopy.py）
            base_info_path = self.process_data_copy()
            
            # 步骤2：生成照片文件名（excel.py）
            self.process_excel(base_info_path)
            
            # 步骤3：生成PDF（emp.py）
            pdf_path = self.generate_pdf()
            
            self.log("===== 生成流程完成 =====")
            messagebox.showinfo("成功", f"PDF文件已生成：\n{pdf_path}")
            
            # 尝试自动打开PDF
            try:
                os.startfile(pdf_path)
            except:
                self.log("提示：无法自动打开PDF，请手动查看")
                
        except Exception as e:
            self.log(f"生成失败：{str(e)}")
            messagebox.showerror("错误", f"生成过程出错：\n{str(e)}")

if __name__ == "__main__":
    # 解决Windows控制台中文乱码（开发时用）
    if sys.platform == 'win32':
        try:
            ctypes.windll.kernel32.SetConsoleCP(65001)
            ctypes.windll.kernel32.SetConsoleOutputCP(65001)
        except:
            pass
    
    root = tk.Tk()
    app = NameplateGeneratorApp(root)
    root.mainloop()
